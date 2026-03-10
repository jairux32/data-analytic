"""
API de Reportes para exportación de datos.

Endpoints para generar y descargar reportes en PDF y Excel.
"""
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_user, require_permission
from app.models.user import User
from app.models.survey import Survey

router = APIRouter(prefix="/api/reports", tags=["reportes"])


@router.get("/excel")
def export_excel(
    province: str = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("exportar_reportes"))
):
    """
    Exporta datos a formato Excel.

    Genera un workbook con múltiples hojas: datos completos,
    resúmenes por dimensión, y coordenadas GPS.
    """
    # Obtener datos
    query = db.query(Survey)
    if province:
        query = query.filter(Survey.province == province)
    surveys = query.all()

    # Crear workbook
    wb = Workbook()

    # Hoja 1: Datos completos
    ws1 = wb.active
    ws1.title = "Datos Completos"

    # Encabezados
    headers = [
        "ID", "UUID", "Productor", "Género", "Edad", "Provincia",
        "Cantón", "Parroquia", "Hectáreas", "Ingreso Total",
        "Ingreso Cacao", "Variedad", "Rendimiento", "Miembro Org.",
    ]
    ws1.append(headers)

    # Datos
    for s in surveys:
        ws1.append([
            s.id,
            s.kobo_uuid,
            s.producer_name,
            s.gender,
            s.age,
            s.province,
            s.canton,
            s.parish,
            s.farm_size_hectares,
            s.total_income,
            s.income_from_cacao,
            s.cacao_varieties,
            s.yield_quintals_per_ha,
            s.member_of_organization,
        ])

    # Hoja 2: Resumen Socioeconómico
    ws2 = wb.create_sheet("Socioeconómico")
    ws2.append(["Campo", "Valor"])
    ws2.append(["Total Productores", len(surveys)])
    if surveys:
        avg_age = sum(s.age for s in surveys if s.age) / len([s for s in surveys if s.age])
        avg_farm = sum(s.farm_size_hectares for s in surveys if s.farm_size_hectares) / len([s for s in surveys if s.farm_size_hectares])
        ws2.append(["Edad Promedio", f"{avg_age:.1f}"])
        ws2.append(["Finca Promedio (ha)", f"{avg_farm:.1f}"])

    # Hoja 3: GPS
    ws3 = wb.create_sheet("Coordenadas GPS")
    ws3.append(["ID", "Productor", "Latitud", "Longitud", "Precisión"])
    for s in surveys:
        if s.gps_latitude and s.gps_longitude:
            ws3.append([s.id, s.producer_name, s.gps_latitude, s.gps_longitude, s.gps_accuracy])

    # Guardar a bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=reporte_cacao_{datetime.now().strftime('%Y%m%d')}.xlsx"
        }
    )


@router.get("/pdf")
def export_pdf(
    province: str = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("exportar_reportes"))
):
    """
    Exporta datos a formato PDF.

    Genera un reporte con resumen ejecutivo, KPIs y tabla de datos.
    """
    # Obtener datos
    query = db.query(Survey)
    if province:
        query = query.filter(Survey.province == province)
    surveys = query.all()

    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Título
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=18,
        textColor=colors.HexColor("#1a5f2a"),
    )
    elements.append(Paragraph("Diagnóstico Socioambiental Cacaotero", title_style))
    elements.append(Spacer(1, 12))

    # Fecha
    elements.append(Paragraph(f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y')}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    # Resumen ejecutivo
    elements.append(Paragraph("Resumen Ejecutivo", styles["Heading2"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f"Total de productores encuestados: {len(surveys)}", styles["Normal"]))

    if surveys:
        provinces = set(s.province for s in surveys if s.province)
        cantons = set(s.canton for s in surveys if s.canton)
        elements.append(Paragraph(f"Provincias cubiertas: {', '.join(provinces)}", styles["Normal"]))
        elements.append(Paragraph(f"Cantones cubiertos: {len(cantons)}", styles["Normal"]))

    elements.append(Spacer(1, 12))

    # Tabla de datos
    elements.append(Paragraph("Productores Encuestados", styles["Heading2"]))
    elements.append(Spacer(1, 6))

    # Crear tabla
    table_data = [["#", "Productor", "Provincia", "Cantón", "Hectáreas"]]
    for i, s in enumerate(surveys[:50], 1):  # Limitar a 50 registros
        table_data.append([
            str(i),
            s.producer_name or "N/A",
            s.province or "N/A",
            s.canton or "N/A",
            f"{s.farm_size_hectares:.1f}" if s.farm_size_hectares else "N/A",
        ])

    if len(surveys) > 50:
        table_data.append(["...", "...", "...", "...", "..."])

    table = Table(table_data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a5f2a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)

    # Construir PDF
    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=reporte_cacao_{datetime.now().strftime('%Y%m%d')}.pdf"
        }
    )
